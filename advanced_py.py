# *****************---------------- Working with Excel Spreadsheets

import openpyxl

# work_book = openpyxl.Workbook()

work_book = openpyxl.load_workbook("users.xlsx")
# print(work_book.sheetnames)

sheet = work_book["Sheet_1"]
# print(sheet)

# --------------------------------------------------------------
# *-*-*-*-*-* Accessing individual cells -----------------------

# Approach one ->>>> *-*- cell coordiantes --------
# cell = sheet["A1"]
# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)

# Approach two ->>>> *-*- cell() method --------
# cell = sheet.cell(row=1, column=1)
# print(cell)

# ----------------------------------------------------------------
# *-*-*-*-*-* Accessing rows/columns cells -----------------------

# print("Number of Rows:", sheet.max_row)
# print("Number of Columns:", sheet.max_column)

# ---------------------------------------------------------------------
# *-*-*-*-*-* Iterating over rows/columns cells -----------------------

# for row in range(1, sheet.max_row + 1):
#     for col in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, col)
#         print(cell.value)

# for col in range(1, sheet.max_column + 1):
#     for row in range(1, sheet.max_row + 1):
#         cell = sheet.cell(row, col)
#         print(cell.value)

# --------------------------------------------------------------
# *-*-*-*-*-* Accessing a range of cells -----------------------


# *-*- getting all the cells in a col --------
# column = sheet["A"]
# print(column)

# *-*- getting a range of cells --------
# print(sheet["A1:B5"])

# *-*- getting a range of cells using coodinates --------
# print(sheet["A:C"])

# *-*- getting all the cells in the first row --------
# print(sheet[1])

# *-*- getting all the cells in a range of rows --------
# print(sheet[1:3])

# --------------------------------------------------------------
# *-*-*-*-*-* sheet object methods -----------------------------

# *-*- append()
# sheet.append(["wolf", 555, "online"])

# *-*- insert_rows(starting index, # of empty rows)
# sheet.insert_rows(3, 10)

# *-*- insert_cols(starting index, # of empty cols)
# sheet.insert_cols(2, 5)

# *-*- delete_rows(starting index, # of rows)
# sheet.delete_rows(3, 10)

# *-*- delete_cols(starting index, # of cols)
# sheet.delete_cols(2, 5)


# work_book.save("users.xlsx")


# --------------------------------------------------------------
# *-*-*-*-*-* adding sheets to the workbook -----------------------------
work_book.create_sheet("PY_Sheet_1")
work_book.save("new_users.xlsx")



# ****************------------------- NumPy

# pipenv install numpy

import numpy as np


# *-*-*-*-*-*-*-*-*-*-*-*- Creating a simple array
# array = np.array([1, 2, 3])
# print(array)
# print(type(array))


# *-*-*-*-*-*-*-*-*-*-*-*- Creating a multidimensional array
# array = np.array([[1, 2, 3, 4], [5, 6, 7, 8]])
# print(array)

# print(array.shape)

# *-*-*-*-*-*-*-*-*-*-*-*- NumPy Helper Methods -------------------------

# -*-*-*- zeros()
# array = np.zeros((4, 6))
# print(array)

# -*-*-*- ones()
# array = np.ones((4, 6), dtype=int)
# print(array)

# -*-*-*- full()
# array = np.full((4, 6), 11, dtype=int)
# print(array)

# -*-*-*- random.random()
# array = np.random.random((3, 2))
# print(array)

# accessing the number at the first row and first col
# print(array[0, 0])

# using a comparison operator
# print(array > 0.5)

# boolean indexing
# print(array[array > 0.3])

# *-*-*-*-*-*-*-*-*-*-*-*- Array Computational Functions ---------------------------

# array = np.array([2.1, 3.9, 5.5, 11.8, 15.7, 13.3])
# print(array)

# -*-*-*- sum()
# print(np.sum(array))

# -*-*-*- floor()
# print(np.floor(array))


# -*-*-*- ceil()
# print(np.ceil(array))


# -*-*-*- round()
# print(np.round(array))


# *-*-*-*-*-*-*-*-*-*-*-*- Performing Arithmetic Operations Between Numbers and Arrays

first_array = np.array([111, 333, 555])
second_array = np.array([222, 444, 666])

print(first_array + second_array)

print(first_array + 1)

# *-*-*-*-*-*-* Unit Conversion (Legnth)
feet = np.array([1, 2, 3, 4, 5])
cm = feet * 30.48
inch = feet * 12
meter = feet * 0.3048
yard = feet * 0.3333
print(cm)
print(inch)
print(meter)
print(yard)
